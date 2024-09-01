import sys, time, pysftp, getpass, subprocess                                       # pysftp for server upload is otional
from openpyxl import *                                                              # openpyxl need to be installed with pip

def cli_progress(start_val:int,end_val:int,bar_length=40):                          # simple progressbar
    percent = float(start_val) / end_val
    hashes = '#' * int(round(percent * bar_length))
    spaces = ' ' * (bar_length - len(hashes))
    if start_val == end_val-1:
        hashes = '#' * bar_length
        spaces = ''
        percent = 1
    sys.stdout.write("\r [{0}] {1}%".format(hashes + spaces, int(round(percent * 100))))
    sys.stdout.flush()

sec = input("\rSeconds to capture: ")
sys.stdout.flush()
print("capturing information --->")
wb = Workbook()
ws = wb.active
ws.title = "Temp, Clock measurement"
ws.cell(row=1, column=1).value = "CPU Package Temp"
ws.cell(row=1, column=2).value = "GPU Package Temp"
ws.cell(row=1, column=3).value = "CPU Clock Speed"

for x in range(1,int(sec)):
    cpu = subprocess.run(["cat", "/sys/class/thermal/thermal_zone0/temp"], capture_output=True, encoding="utf-8").stdout.strip()
    gpu = subprocess.run(["cat", "/sys/class/thermal/thermal_zone1/temp"], capture_output=True, encoding="utf-8").stdout.strip()
    clk = subprocess.run(["cat", "/sys/devices/system/cpu/cpu0/cpufreq/scaling_cur_freq"], capture_output=True, encoding="utf-8").stdout.strip()
    ws.cell(row=x+1, column=1).value = float(round(int(cpu) / 1000, 1))
    ws.cell(row=x+1, column=2).value = float(round(int(gpu) / 1000, 1))
    ws.cell(row=x+1, column=3).value = float(round(int(clk) / 1000, 1))
    cli_progress(x, int(sec), 60)
    time.sleep(1)

filename = input("\nEnter Filename: ")
wb.save(f'{filename}.xlsx')
print("File saved to current directory")

#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
cnopts = pysftp.CnOpts()                                                                                                        # Comment out this part if server upload is not required
cnopts.hostkeys = None
print("Credentials for SFTP to Droppy\n")
with pysftp.Connection('IPADDRESS', username=input("Username: "), password=getpass.getpass(), cnopts=cnopts) as sftp:           # Replace IPADRESS with desired IP
    with sftp.cd('SERVERPATH'):                                                                                             	# Replace SERVERPATH with absolute path you want to save to
        sftp.put(f'{filename}.xlsx', preserve_mtime=True)
sftp.close()
print("File uploaded to server")
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------